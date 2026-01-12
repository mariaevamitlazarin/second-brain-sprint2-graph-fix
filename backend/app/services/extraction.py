import csv
import re
from pathlib import Path
from typing import List, Tuple

from openpyxl import load_workbook
from pypdf import PdfReader

from app.core.config import settings
from app.models.document import DocumentRecord, DocumentStatus, ExtractionMetadata
from app.services.registry import registry


class ExtractionResult:
    def __init__(self, text: str, metadata: ExtractionMetadata):
        self.text = text
        self.metadata = metadata


def normalize_text(text: str) -> str:
    text = text.replace("\x00", " ")
    text = re.sub(r"\r\n?", "\n", text)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def count_words(text: str) -> int:
    return len(re.findall(r"\b\w+\b", text, flags=re.UNICODE))


def naive_summary(text: str, max_sentences: int = 3) -> str:
    clean = normalize_text(text)
    if not clean:
        return "No readable text was extracted from this document."

    sentences = re.split(r"(?<=[.!?])\s+", clean)
    selected = [sentence.strip() for sentence in sentences if len(sentence.strip()) > 40]
    if not selected:
        return clean[:500]
    return " ".join(selected[:max_sentences])[:900]


def naive_keywords(text: str, limit: int = 12) -> List[str]:
    stopwords = {
        "about", "after", "again", "also", "among", "because", "before", "being", "between",
        "could", "during", "from", "have", "into", "more", "most", "other", "over", "such",
        "than", "that", "their", "there", "these", "this", "through", "using", "were", "with",
        "would", "para", "como", "com", "das", "dos", "uma", "por", "que", "não", "mais",
        "são", "the", "and", "for", "are", "was", "his", "her", "its", "you", "your", "our"
    }
    words = re.findall(r"\b[a-zA-ZÀ-ÿ][a-zA-ZÀ-ÿ0-9_-]{3,}\b", text.lower())
    counts = {}
    for word in words:
        if word in stopwords:
            continue
        counts[word] = counts.get(word, 0) + 1
    ranked = sorted(counts.items(), key=lambda item: item[1], reverse=True)
    return [word for word, _ in ranked[:limit]]


def extract_text_file(path: Path) -> ExtractionResult:
    text = path.read_text(encoding="utf-8", errors="ignore")
    text = normalize_text(text)
    metadata = ExtractionMetadata(
        text_length=len(text),
        word_count=count_words(text),
        extraction_engine="python-text-reader",
    )
    return ExtractionResult(text, metadata)


def extract_pdf(path: Path) -> ExtractionResult:
    reader = PdfReader(str(path))
    pages = []
    for index, page in enumerate(reader.pages, start=1):
        page_text = page.extract_text() or ""
        if page_text.strip():
            pages.append(f"\n\n--- Page {index} ---\n{page_text}")
    text = normalize_text("\n".join(pages))
    metadata = ExtractionMetadata(
        text_length=len(text),
        word_count=count_words(text),
        page_count=len(reader.pages),
        extraction_engine="pypdf",
    )
    return ExtractionResult(text, metadata)


def extract_csv(path: Path) -> ExtractionResult:
    rows = []
    with path.open("r", encoding="utf-8", errors="ignore", newline="") as file_obj:
        sample = file_obj.read(4096)
        file_obj.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(file_obj, dialect)
        for row in reader:
            rows.append(row)

    text_lines = [" | ".join(cell.strip() for cell in row) for row in rows]
    column_count = max((len(row) for row in rows), default=0)
    text = normalize_text("\n".join(text_lines))
    metadata = ExtractionMetadata(
        text_length=len(text),
        word_count=count_words(text),
        row_count=len(rows),
        column_count=column_count,
        extraction_engine="python-csv",
    )
    return ExtractionResult(text, metadata)


def extract_xlsx(path: Path) -> ExtractionResult:
    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    text_lines = []
    total_rows = 0
    max_columns = 0
    sheet_names = workbook.sheetnames

    for sheet in workbook.worksheets:
        text_lines.append(f"\n--- Sheet: {sheet.title} ---")
        for row in sheet.iter_rows(values_only=True):
            values = ["" if cell is None else str(cell) for cell in row]
            if any(value.strip() for value in values):
                text_lines.append(" | ".join(values))
                total_rows += 1
                max_columns = max(max_columns, len(values))

    text = normalize_text("\n".join(text_lines))
    metadata = ExtractionMetadata(
        text_length=len(text),
        word_count=count_words(text),
        row_count=total_rows,
        column_count=max_columns,
        sheet_names=sheet_names,
        extraction_engine="openpyxl",
    )
    return ExtractionResult(text, metadata)


def extract_xls(path: Path) -> ExtractionResult:
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError("xlrd is required to process .xls files.") from exc

    workbook = xlrd.open_workbook(str(path))
    text_lines = []
    total_rows = 0
    max_columns = 0
    sheet_names = workbook.sheet_names()

    for sheet in workbook.sheets():
        text_lines.append(f"\n--- Sheet: {sheet.name} ---")
        for row_index in range(sheet.nrows):
            values = [str(sheet.cell_value(row_index, col_index)) for col_index in range(sheet.ncols)]
            if any(value.strip() for value in values):
                text_lines.append(" | ".join(values))
                total_rows += 1
                max_columns = max(max_columns, len(values))

    text = normalize_text("\n".join(text_lines))
    metadata = ExtractionMetadata(
        text_length=len(text),
        word_count=count_words(text),
        row_count=total_rows,
        column_count=max_columns,
        sheet_names=sheet_names,
        extraction_engine="xlrd",
    )
    return ExtractionResult(text, metadata)


def extract_by_extension(path: Path, extension: str) -> ExtractionResult:
    extension = extension.lower()
    if extension in {".md", ".txt"}:
        return extract_text_file(path)
    if extension == ".pdf":
        return extract_pdf(path)
    if extension == ".csv":
        return extract_csv(path)
    if extension == ".xlsx":
        return extract_xlsx(path)
    if extension == ".xls":
        return extract_xls(path)
    raise ValueError(f"Unsupported file type: {extension}")


def process_document(document: DocumentRecord) -> DocumentRecord:
    if document.duplicate_of is not None:
        document.status = DocumentStatus.duplicate
        registry.update(document)
        return document

    if not document.stored_filename:
        document.status = DocumentStatus.failed
        document.extraction_error = "Document has no stored file."
        registry.update(document)
        return document

    source_path = settings.upload_dir / document.stored_filename
    if not source_path.exists():
        document.status = DocumentStatus.failed
        document.extraction_error = f"Stored file not found: {document.stored_filename}"
        registry.update(document)
        return document

    document.status = DocumentStatus.processing
    document.extraction_error = None
    registry.update(document)

    try:
        result = extract_by_extension(source_path, document.file_extension)
        extracted_filename = f"{document.id}.txt"
        extracted_path = settings.extracted_dir / extracted_filename
        extracted_path.write_text(result.text, encoding="utf-8")

        document.extracted_text_path = extracted_filename
        document.text_excerpt = result.text[: settings.text_excerpt_chars]
        document.summary = naive_summary(result.text)
        document.keywords = naive_keywords(result.text)
        document.extraction_metadata = result.metadata
        document.status = DocumentStatus.processed
        document.extraction_error = None
    except Exception as exc:
        document.status = DocumentStatus.failed
        document.extraction_error = str(exc)

    registry.update(document)
    return document


def read_extracted_text(document: DocumentRecord) -> str:
    if not document.extracted_text_path:
        return ""
    path = settings.extracted_dir / document.extracted_text_path
    if not path.exists():
        return ""
    return path.read_text(encoding="utf-8", errors="ignore")
